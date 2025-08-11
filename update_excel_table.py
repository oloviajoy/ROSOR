import os
import re
import datetime
import xlsxwriter
import pandas as pd
import logging
import sys
import math
from typing import List, Tuple
import json
import xml.etree.ElementTree as ET
from openpyxl import load_workbook as openpyxl_load_workbook


# Setup logging for debugging
logging.basicConfig(level=logging.DEBUG, format='[DEBUG] %(message)s')

# Configuration
input_folder      = r"Y:\TECK_WHITE_EARTH\TECK_HYPERSPEC"
flight_folder_2D  = r"Z:\Rosor\2024 - 2025\Projects\2502-58-3-dlmrh-teck-white-earth\4.Acquisition\Hyspec_flts"
base_output       = os.path.join(input_folder, "teck_folders.xlsx")
match_threshold_seconds = 30

FILE_ALLCUBES   = "_AllCubes.tif"
FILE_WHITEREF   = "sceneWhiteReference.hdr"


def make_unique_filename(full_path):
    dir_name, base = os.path.split(full_path)
    name, ext       = os.path.splitext(base)
    search_dir      = dir_name or os.getcwd()
    max_v = 0
    for f in os.listdir(search_dir):
        m = re.match(rf'^{re.escape(name)}_v(\d+){re.escape(ext)}$', f)
        if m:
            max_v = max(max_v, int(m.group(1)))
    if max_v > 0:
        new_base = f"{name}_v{max_v+1}{ext}"
    else:
        if os.path.exists(os.path.join(search_dir, base)):
            new_base = f"{name}_v2{ext}"
        else:
            new_base = base
    logging.debug(f"Unique filename resolved: {new_base}")
    return os.path.join(search_dir, new_base)



def count_flight_lines(kml_path: str) -> int:
    """
    Given a full path to a 2D-flight KML, parse its <description> CDATA
    as JSON and return the number of items in the 'children' array.
    Returns 0 on any parse error or if no children are found.
    """
    try:
        tree = ET.parse(kml_path)
        ns = {'kml': 'http://www.opengis.net/kml/2.2'}
        desc_elem = tree.find('.//kml:description', ns)
        if desc_elem is None or not desc_elem.text:
            return 0
        data = json.loads(desc_elem.text)
        return len(data.get('children', []))
    except Exception as e:
        logging.debug(f"Failed to count lines in {kml_path}: {e}")
        return 0


def find_latest_file(full_path):
    dir_name, base = os.path.split(full_path)
    name, ext      = os.path.splitext(base)
    search_dir     = dir_name or os.getcwd()
    versions = []
    for f in os.listdir(search_dir):
        if f == base:
            versions.append((0, f))
        else:
            m = re.match(rf'^{re.escape(name)}_v(\d+){re.escape(ext)}$', f)
            if m:
                versions.append((int(m.group(1)), f))
    if not versions:
        logging.debug(f"No versions found for {base}")
        return None
    latest = max(versions, key=lambda x: x[0])[1]
    latest_path = os.path.join(search_dir, latest)
    logging.debug(f"Latest file found: {latest_path}")
    return latest_path


def natural_sort_key(s):
    return [int(text) if text.isdigit() else text.lower() for text in re.split(r'(\d+)', s)]


def get_list_of_teck_folders(folder_path, prefix="TECK_T"):
    matches = []
    for root, dirs, _ in os.walk(folder_path):
        if "dark_" in root or any(part.lower() == "old" for part in root.split(os.sep)):
            dirs[:] = []
            continue
        dirs[:] = [d for d in dirs if d.lower() != "old"]
        for d in dirs:
            if d.startswith(prefix):
                matches.append(os.path.join(root, d))
        dirs[:] = [d for d in dirs if not d.startswith(prefix)]
    sorted_list = sorted(matches, key=natural_sort_key)
    logging.debug(f"Found {len(sorted_list)} VNIR folders")
    return sorted_list


def _parse_folder_timestamp(name: str) -> datetime.datetime | None:
    m = re.search(r'_(\d{4}_\d{2}_\d{2}_\d{2}_\d{2}_\d{2}(?:_\d{1,6})?)$', name)
    if not m:
        return None
    ts = m.group(1)
    parts = ts.split('_')
    year, mon, day, hr, mi, sec = map(int, parts[:6])
    us = int(parts[6].ljust(6, '0')) if len(parts) == 7 else 0
    return datetime.datetime(year, mon, day, hr, mi, sec, us)


def build_swir_index(base_folder: str) -> dict[str, list[tuple[datetime.datetime, str]]]:
    swir_index: dict[str, list[tuple[datetime.datetime, str]]] = {}
    for root, dirs, _ in os.walk(base_folder):
        if "dark_" in root or any(p.lower() == "old" for p in root.split(os.sep)):
            dirs[:] = []
            continue
        for d in dirs:
            m = re.match(r'^(\d{6}_.+?_\d{4}_\d{2}_\d{2}_\d{2}_\d{2}_\d{2}(?:_\d{1,6})?)$', d)
            if not m:
                continue
            dt = _parse_folder_timestamp(d)
            if dt is None:
                continue
            pm = re.match(r'^\d{6}_(.+?)(?=_\d{4})', d)
            if not pm:
                continue
            prefix = pm.group(1)
            swir_index.setdefault(prefix, []).append((dt, os.path.join(root, d)))
    for lst in swir_index.values():
        lst.sort(key=lambda x: x[0])
    logging.debug(f"Built SWIR index with {len(swir_index)} prefixes")
    return swir_index


def build_kml_index(base_folder: str) -> dict[tuple[str, str], str]:
    pattern = re.compile(r"tof_(\d+)_flt_([0-9n]+)_.*\.kml$", re.IGNORECASE)
    idx: dict[tuple[str, str], str] = {}

    for root, _, files in os.walk(base_folder):
        for f in files:
            if not f.lower().endswith('.kml'):
                continue
            m = pattern.match(f)
            if not m:
                continue
            tof     = m.group(1)
            flights = m.group(2).split('n')
            full    = os.path.join(root, f)
            for fl in flights:
                idx[(tof, fl)] = full

    logging.debug(f"Built KML index with {len(idx)} entries")
    return idx


def find_corresponding_swir(vnir_path: str,
                           swir_index: dict[str, list[tuple[datetime.datetime, str]]],
                           match_threshold_seconds: float = 30
                           ) -> str | None:
    vnir_name = os.path.basename(vnir_path)
    vnir_dt   = _parse_folder_timestamp(vnir_name)
    if vnir_dt is None:
        return None
    m = re.match(r'(.+?)(?=_\d{4})', vnir_name)
    if not m:
        return None
    prefix    = m.group(1)
    candidates = swir_index.get(prefix)
    if not candidates:
        return None
    nearest_dt, nearest_path = min(
        candidates,
        key=lambda tup: abs((tup[0] - vnir_dt).total_seconds())
    )
    if abs((nearest_dt - vnir_dt).total_seconds()) <= match_threshold_seconds:
        return nearest_path
    return None


def extract_takeoff_position(name):
    m = re.match(r"TECK_T(\d+)", name)
    return m.group(1) if m else ""


def extract_flight_numbers(name):
    return re.findall(r"_F(\d+)", name) or [""]


def extract_utc_datetime(name):
    m = re.search(r"_(\d{4})_(\d{2})_(\d{2})_(\d{2})_(\d{2})_(\d{2})_", name)
    if not m:
        return None
    year, mon, day, hr, mi, sec = map(int, m.groups())
    return datetime.datetime(year, mon, day, hr, mi, sec)


def find_date_folder_code(path):
    p = os.path.abspath(path)
    while True:
        p, tail = os.path.split(p)
        if not tail:
            return None
        if re.fullmatch(r"\d{4}", tail):
            return tail



# 1) Build SWIR index
#    (unchanged)
swir_index = build_swir_index(input_folder)

# 1.5) Build 2D KML index entries
logging.debug("Building 2D KML index…")
kml_index = build_kml_index(flight_folder_2D)
logging.debug(f"Built KML index with {len(kml_index)} entries")

# 2) Read previous Excel for human‑editable columns AND for col‑settings
latest_excel = find_latest_file(base_output)
prev_df = None
editable_cols: list[str] = []
prev_col_settings: dict[str, dict] = {}

if latest_excel and os.path.exists(latest_excel):
    try:
        prev_df = pd.read_excel(latest_excel,
                                engine='openpyxl',
                                dtype={'Flight': str})
        prev_df.columns = prev_df.columns.str.strip()
        prev_df['Flight'] = prev_df['Flight']\
                                .str.replace(r'\.0$', '',
                                             regex=True)\
                                .str.strip()
        prev_df['VNIR flight folder name'] = prev_df['VNIR flight folder name']\
                                                .astype(str)\
                                                .str.strip()
        editable_cols = [c for c in prev_df.columns if c.startswith("[E]")]
    except Exception as e:
        raise RuntimeError(f"Previous Excel file exists but cannot be read: {e}")
if "[E] Notes" not in editable_cols:
    editable_cols.append("[E] Notes")

    # --- NEW: read column width + hidden flags from the old sheet ---
    wb_prev = openpyxl_load_workbook(latest_excel, data_only=True)
    ws_prev = wb_prev.active
    # build a map: header → { width, hidden }
    for col_letter, col_dim in ws_prev.column_dimensions.items():
        header = ws_prev[f"{col_letter}1"].value
        if header:
            prev_col_settings[header] = {
                "width": col_dim.width,
                "hidden": bool(col_dim.hidden)
            }

# 3) Determine new output filename
output_excel = make_unique_filename(base_output)

# 4) Collect VNIR folders
vnir_folders = get_list_of_teck_folders(input_folder)

# 5) Create workbook & worksheet
workbook  = xlsxwriter.Workbook(output_excel)
worksheet = workbook.add_worksheet()

# 6) Define formats
dt_fmt   = workbook.add_format({'num_format': 'yyyy-mm-dd hh:mm:ss'})
date_fmt = workbook.add_format({'num_format': 'yyyy-mm-dd'})
header_fmt = workbook.add_format({'bold': True, 'bg_color': '#D7E4BC', 'border': 1})

# 7) Headers (include SWIR Dark folder check)
base_headers = [
    'VNIR flight folder name', 'Open VNIR Folder',
    'Corresponding SWIR', 'Open SWIR Folder',
    'Over-exposure check', 'Tarp check',
    'SWIR WhiteRef check', 'SWIR Dark folder check',  # new column
    'Take-Off', 'Flight', 'Combined flt count',
    'UTC Date', 'Date Folder',
    FILE_ALLCUBES, '2D KML',
    '# of flight lines'
]
headers = base_headers + editable_cols
worksheet.write_row('A1', headers, header_fmt)

# Determine column indexes
vnir_url_col     = headers.index('Open VNIR Folder')
swir_url_col     = headers.index('Open SWIR Folder')
swir_whref_col   = headers.index('SWIR WhiteRef check')
swir_dark_col    = headers.index('SWIR Dark folder check')
allcubes_col_idx = headers.index(FILE_ALLCUBES)
kml_col_idx      = headers.index('2D KML')
flight_lines_col_idx  = headers.index('# of flight lines')

# 8) Prepare carry-over map
prev_map = {}
if prev_df is not None:
    for _, prow in prev_df.iterrows():
        key = f"{prow['VNIR flight folder name']}|{prow['Flight']}"
        prev_map[key] = {col: prow.get(col, '') for col in editable_cols}
    logging.debug(f"Prepared prev_map with {len(prev_map)} entries")

# 9) Track max widths
max_widths = [len(h) for h in headers]

# 10) Populate rows...
logging.debug("Populating rows...")
row = 1
for idx, vnir_path in enumerate(vnir_folders):
    vnir_name = os.path.basename(vnir_path)
    takeoff   = extract_takeoff_position(vnir_name)
    flights   = extract_flight_numbers(vnir_name)
    code      = find_date_folder_code(vnir_path)
    utc_dt    = extract_utc_datetime(vnir_name)
    date_folder_dt = None
    if code and utc_dt:
        mon, day = int(code[:2]), int(code[2:])
        date_folder_dt = datetime.date(utc_dt.year, mon, day)

    # find corresponding SWIR
    swir_path  = find_corresponding_swir(vnir_path, swir_index, match_threshold_seconds)
    swir_name  = os.path.basename(swir_path) if swir_path else ''

    # VNIR checks\
    over_exp = '✅' if os.path.exists(os.path.join(vnir_path, FILE_ALLCUBES)) else '❌'
    tarp     = '✅' if os.path.exists(os.path.join(vnir_path, FILE_WHITEREF)) else '❌'

    # AllCubes column
    allcubes_exists = os.path.exists(os.path.join(vnir_path, FILE_ALLCUBES))
    allcubes_file   = FILE_ALLCUBES if allcubes_exists else ''
    allcubes_path   = os.path.join(vnir_path, FILE_ALLCUBES) if allcubes_exists else ''

    # SWIR WhiteRef check
    swir_whref_exists = bool(swir_path and os.path.exists(os.path.join(swir_path, FILE_WHITEREF)))
    swir_whref        = '✅' if swir_whref_exists else '❌'

    # SWIR Dark folder check
    swir_dark_exists = False
    if swir_path:
        for sub in os.listdir(swir_path):
            sub_path = os.path.join(swir_path, sub)
            if os.path.isdir(sub_path) and 'dark' in sub.lower():
                swir_dark_exists = True
                break
    swir_dark = '✅' if swir_dark_exists else '❌'

    logging.debug(
        f"Folder[{idx}]: {vnir_name} swir={swir_name} over_exp={over_exp} "
        f"tarp={tarp} swir_whref={swir_whref} swir_dark={swir_dark} allcubes={allcubes_file}"
    )

    for combined_count, flight in enumerate(flights, start=1):
        key   = f"{vnir_name}|{flight}"
        notes = prev_map.get(key, {})
        kml_path = kml_index.get((takeoff, flight), '')
        line_count = count_flight_lines(kml_path) if kml_path else ''

        vals = [
            vnir_name,
            "Open Folder",
            swir_name,
            "Open Folder" if swir_path else '',
            over_exp,
            tarp,
            swir_whref,
            swir_dark,                        # new column value
            takeoff,
            flight,
            combined_count,
            utc_dt.strftime("%Y-%m-%d %H:%M:%S") if utc_dt else '',
            date_folder_dt.strftime("%Y-%m-%d") if date_folder_dt else '',
            allcubes_file,
            os.path.basename(kml_path) if kml_path else '',  # the “2D KML” column
            line_count  # the new “# of flight lines” column
        ]
        vals += [notes.get(col, '') for col in editable_cols]

        for col_idx, val in enumerate(vals):
            if isinstance(val, float) and math.isnan(val):
                val = ''
            if col_idx == vnir_url_col:
                worksheet.write_url(row, col_idx, f"file:///{vnir_path}", string=val)
            elif col_idx == swir_url_col and swir_path:
                worksheet.write_url(row, col_idx, f"file:///{swir_path}", string=val)
            elif col_idx == swir_dark_col:
                worksheet.write(row, col_idx, swir_dark)
            elif col_idx == allcubes_col_idx and allcubes_path:
                worksheet.write_url(row, col_idx, f"file:///{allcubes_path}", string=val)
            elif col_idx == kml_col_idx and kml_path:
                worksheet.write_url(row, col_idx, f"file:///{kml_path}", string="Open KML")
            elif col_idx == 11 and utc_dt:
                worksheet.write_datetime(row, col_idx, utc_dt, dt_fmt)
            elif col_idx == 12 and date_folder_dt:
                worksheet.write_datetime(row, col_idx, date_folder_dt, date_fmt)
            else:
                worksheet.write(row, col_idx, val)

        for col_idx, text in enumerate(vals):
            max_widths[col_idx] = max(max_widths[col_idx], len(str(text)))

        logging.debug(f"Wrote row {row} with key={key}")
        row += 1

# 11) Add table
worksheet.add_table(0, 0, row - 1, len(headers) - 1, {
    'name':  'TECKFlights',
    'style': 'Table Style Medium 9',
    'columns': [{'header': h} for h in headers]
})

# 12) Apply old widths/hidden or fall back on auto‑width
def xl_col_to_name(col_idx: int) -> str:
    """Convert 0‑based idx → Excel column name."""
    name = ""
    while col_idx >= 0:
        name = chr(col_idx % 26 + ord('A')) + name
        col_idx = col_idx // 26 - 1
    return name

for col_idx, header in enumerate(headers):
    letter = xl_col_to_name(col_idx)
    settings = prev_col_settings.get(header, {})
    # use previous width if present, else auto‑width + padding
    width = settings.get("width", max_widths[col_idx] + 2)
    hidden = settings.get("hidden", False)
    # xlsxwriter: pass hidden via the options dict
    worksheet.set_column(f"{letter}:{letter}",
                         width,
                         None,
                         {'hidden': hidden})

# 13) Save and open
workbook.close()
try:
    os.startfile(output_excel)
except Exception:
    logging.warning("Could not auto-open the Excel file.")